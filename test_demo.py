from selenium import webdriver
from selenium.webdriver.common.by import By
from time import sleep
from selenium.webdriver.support.wait import WebDriverWait #ilgili driverı bekleten yapı
from selenium.webdriver.support import expected_conditions as ec #beklenen koşullar
from selenium.webdriver.common.action_chains import ActionChains 
import pytest
import openpyxl

class Test_Demo:
    def deneme():
        print("deneme")

    #pytest tarafından tanımlanan bir method 
    #her test öncesi otomatik olarak çalıştırılır
    def setup_method(self):
        self.driver = webdriver.Chrome()
        self.driver.maximize_window()
        self.driver.get("https://www.saucedemo.com/")

    #her test bitiminde çalışacak fonk
    def teardown_method(self):
        self.driver.quit()

    @pytest.mark.skip #tüm testler koşulurken "skip" şeklinde işaretlenen testlerimi atla
    def test_demo(self):
        print("x")
        text = "Hello"
        assert text == "Hello"

    def getData():
        return [("1","1"),("abc","123"),("deneme","secret_sauce")]
    
    def readInvalidDataFromExcel():
        excelFile = openpyxl.load_workbook("data/invalidLogin.xlsx")
        sheet = excelFile["Sheet1"]
        rows = sheet.max_row #kaçıncı satıra kadar benim verim var
        data = []
        for i in range(2,rows+1):
            username = sheet.cell(i,1).value
            password = sheet.cell(i,2).value
            data.append((username,password))
        return data

    
    @pytest.mark.parametrize("username,password",readInvalidDataFromExcel())
    def test_invalid_login(self,username,password):
        userNameInput = WebDriverWait(self.driver,5).until(ec.visibility_of_element_located((By.ID,"user-name")))
        passwordInput = WebDriverWait(self.driver,5).until(ec.visibility_of_element_located((By.ID,"password")))
        userNameInput.send_keys(username)
        passwordInput.send_keys(password)
        loginButton = WebDriverWait(self.driver,5).until(ec.visibility_of_element_located((By.ID,"login-button")))
        loginButton.click()
        errorMessage =WebDriverWait(self.driver,5).until(ec.visibility_of_element_located((By.XPATH,"//*[@id='login_button_container']/div/form/div[3]/h3")))
        assert errorMessage.text == "Epic sadface: Username and password do not match any user in this service"

    def test_valid_login(self):
        self.driver.get("https://www.saucedemo.com/")
        userNameInput = WebDriverWait(self.driver,5).until(ec.visibility_of_element_located((By.ID,"user-name")))
        passwordInput =WebDriverWait(self.driver,5).until(ec.visibility_of_element_located((By.ID,"password")))
        actions = ActionChains(self.driver)
        actions.send_keys_to_element(userNameInput,"standard_user")
        actions.send_keys_to_element(passwordInput,"secret_sauce")
        actions.perform() #depoladığım aksiyonları çalıştır
        loginButton = WebDriverWait(self.driver,5).until(ec.visibility_of_element_located((By.ID,"login-button")))
        loginButton.click()
        baslik =WebDriverWait(self.driver,5).until(ec.visibility_of_element_located((By.XPATH,"//*[@id='header_container']/div[1]/div[2]/div")))
        assert baslik.text == "Swag Labs"
        
        